import sys, json
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
wb = openpyxl.load_workbook(r'C:\Users\Admin\Desktop\TEST WEB.xlsx')
ws = wb.active

results = []
for row in range(2, ws.max_row+1):
    mail = ws.cell(row, 2).value
    if not mail or str(mail).strip() == '':
        continue
    stt = ws.cell(row, 1).value
    pw = ws.cell(row, 3).value
    mail_kp = ws.cell(row, 4).value
    twofa = ws.cell(row, 5).value
    sdt = ws.cell(row, 6).value
    link_otp = ws.cell(row, 7).value
    
    idx = len(results)
    obj = {
        "id": 1001 + idx,
        "email": str(mail).strip(),
        "pass": str(pw).strip() if pw else "",
        "recovery": str(mail_kp).strip() if mail_kp else "",
        "type": "SATELLITE",
        "status": "LIVE",
        "workStatus": "Chua lam",
        "channelStatus": "",
        "twoFA": str(twofa).strip() if twofa else "",
        "phone": str(sdt).strip() if sdt else "",
        "otpLink": str(link_otp).strip() if link_otp else "",
        "links": [],
        "createdAt": "2024-05-11",
        "assigneeId": "",
        "assignedTo": "",
        "batchName": ""
    }
    results.append(obj)

print(f'Total: {len(results)}')
# Output as TS array
lines = []
for r in results:
    lines.append(f'''  {{
    id: {r["id"]},
    email: "{r["email"]}",
    pass: "{r["pass"]}",
    recovery: "{r["recovery"]}",
    type: "SATELLITE" as const,
    status: "LIVE" as const,
    workStatus: "Ch\\u01b0a l\\u00e0m",
    channelStatus: "",
    twoFA: "{r["twoFA"]}",
    phone: "{r["phone"]}",
    otpLink: "{r["otpLink"]}",
    links: [],
    createdAt: "2024-05-11",
    assigneeId: "",
    assignedTo: "",
    batchName: ""
  }}''')

ts_code = ',\n'.join(lines)
with open('satellite_data.txt', 'w', encoding='utf-8') as f:
    f.write(ts_code)
print('Written to satellite_data.txt')
